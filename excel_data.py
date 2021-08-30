from openpyxl import load_workbook


class ExcelLoad:
    def __init__(self, excel_path: str):
        # 加载excel文件
        self.wb = None
        try:
            self.wb = load_workbook(excel_path)
        except:
            print('文件加载失败')
            raise
        self.sh = None

    def write_data(self, sheet_name):
        # 传入表格sheet名称 获取当表格
        self.sh = self.wb[sheet_name]
        return self.sh
        # row = 2


if __name__ == '__main__':
    exl = ExcelLoad(excel_path=r'C:\Users\Administrator\Documents\GitHub\BaiduTranslate\百度翻译.xlsx')
    sheet = exl.write_data('Sheet1')
    print(sheet.cell(1, 1).value)
    # 获取表格的总行数
    print(sheet.max_row)
